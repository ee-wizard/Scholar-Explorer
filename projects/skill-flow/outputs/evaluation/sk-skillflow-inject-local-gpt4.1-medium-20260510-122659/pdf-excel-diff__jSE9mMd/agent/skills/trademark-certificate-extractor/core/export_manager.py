#!/usr/bin/env python3
"""
Export Manager Module

Manages multiple export formats (Excel, JSON, CSV, PDF reports).
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
from enum import Enum


class ExportFormat(Enum):
    """Supported export formats"""
    EXCEL = "excel"
    JSON = "json"
    CSV = "csv"
    PDF_REPORT = "pdf_report"
    HTML_REPORT = "html_report"


class ExportManager:
    """
    Manages export of trademark data to multiple formats
    """

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize export manager

        Args:
            output_dir: Directory for output files (defaults to current directory)
        """
        self.output_dir = output_dir or Path.cwd()
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export(self, data: List[Dict[str, Any]],
               format: ExportFormat,
               filename: Optional[str] = None) -> Path:
        """
        Export data to specified format

        Args:
            data: List of trademark dictionaries
            format: Export format
            filename: Optional filename (auto-generated if None)

        Returns:
            Path to exported file
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"trademark_export_{timestamp}"

        if format == ExportFormat.EXCEL:
            return self.export_to_excel(data, filename)
        elif format == ExportFormat.JSON:
            return self.export_to_json(data, filename)
        elif format == ExportFormat.CSV:
            return self.export_to_csv(data, filename)
        elif format == ExportFormat.PDF_REPORT:
            return self.export_to_pdf_report(data, filename)
        elif format == ExportFormat.HTML_REPORT:
            return self.export_to_html_report(data, filename)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def export_to_excel(self, data: List[Dict[str, Any]],
                       filename: str) -> Path:
        """Export to Excel format (delegates to existing generate_excel.py)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as OpenpyxlImage
        except ImportError:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")

        output_path = self.output_dir / f"{filename}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "商标证书信息"

        # Headers
        headers = ['序号', '商标标识', '注册人', '注册号', '国际分类', '有效期限',
                  '置信度', '到期状态', '需要复核']

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for row_num, entry in enumerate(data, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=3, value=entry.get('注册人', ''))
            ws.cell(row=row_num, column=4, value=entry.get('注册号', ''))
            ws.cell(row=row_num, column=5, value=entry.get('国际分类', ''))
            ws.cell(row=row_num, column=6, value=entry.get('有效期限', ''))

            # Confidence score
            confidence = entry.get('overall_confidence', 0)
            ws.cell(row=row_num, column=7, value=f"{confidence:.1%}" if confidence else "N/A")

            # Expiry status
            ws.cell(row=row_num, column=8, value=entry.get('expiry_message', ''))

            # Review flag
            needs_review = entry.get('requires_manual_review', False)
            ws.cell(row=row_num, column=9, value="是" if needs_review else "否")

            # Add logo if available
            logo_path_str = entry.get('商标标识图片路径', '')
            if logo_path_str:
                logo_path = Path(logo_path_str)
                if logo_path.exists():
                    try:
                        img = OpenpyxlImage(str(logo_path))
                        img.height = 50
                        img.width = 150
                        img.anchor = f"B{row_num}"
                        ws.add_image(img)
                        ws.row_dimensions[row_num].height = 60
                    except:
                        pass

        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        ws.column_dimensions['B'].width = 25  # Logo column wider

        wb.save(output_path)
        return output_path

    def export_to_json(self, data: List[Dict[str, Any]],
                      filename: str) -> Path:
        """Export to JSON format"""
        output_path = self.output_dir / f"{filename}.json"

        # Clean data for JSON serialization
        clean_data = []
        for item in data:
            clean_item = {}
            for key, value in item.items():
                # Skip non-serializable items
                if isinstance(value, (str, int, float, bool, list, dict, type(None))):
                    clean_item[key] = value
                elif hasattr(value, 'isoformat'):  # datetime objects
                    clean_item[key] = value.isoformat()
                else:
                    clean_item[key] = str(value)
            clean_data.append(clean_item)

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(clean_data, f, ensure_ascii=False, indent=2)

        return output_path

    def export_to_csv(self, data: List[Dict[str, Any]],
                     filename: str) -> Path:
        """Export to CSV format"""
        output_path = self.output_dir / f"{filename}.csv"

        if not data:
            # Create empty file
            with open(output_path, 'w', encoding='utf-8-sig') as f:
                f.write("")
            return output_path

        # Determine fields to export
        fieldnames = ['序号', '注册人', '注册号', '国际分类', '有效期限',
                     '置信度', '到期状态', '需要复核']

        with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()

            for i, item in enumerate(data, 1):
                row = {
                    '序号': i,
                    '注册人': item.get('注册人', ''),
                    '注册号': item.get('注册号', ''),
                    '国际分类': item.get('国际分类', ''),
                    '有效期限': item.get('有效期限', ''),
                    '置信度': f"{item.get('overall_confidence', 0):.1%}" if item.get('overall_confidence') else 'N/A',
                    '到期状态': item.get('expiry_message', ''),
                    '需要复核': '是' if item.get('requires_manual_review') else '否',
                }
                writer.writerow(row)

        return output_path

    def export_to_html_report(self, data: List[Dict[str, Any]],
                             filename: str) -> Path:
        """Export to HTML report format"""
        output_path = self.output_dir / f"{filename}.html"

        html_content = self._generate_html_report(data)

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return output_path

    def _generate_html_report(self, data: List[Dict[str, Any]]) -> str:
        """Generate HTML report content"""
        # Calculate statistics
        total = len(data)
        high_confidence = sum(1 for d in data if d.get('confidence_level') == 'high')
        needs_review = sum(1 for d in data if d.get('requires_manual_review'))
        expiring_soon = sum(1 for d in data if d.get('requires_renewal'))

        html = f"""
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>商标证书提取报告</title>
    <style>
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            border-bottom: 3px solid #4472C4;
            padding-bottom: 10px;
        }}
        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 30px 0;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }}
        .stat-card.green {{
            background: linear-gradient(135deg, #56ab2f 0%, #a8e063 100%);
        }}
        .stat-card.orange {{
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        }}
        .stat-card.red {{
            background: linear-gradient(135deg, #fa709a 0%, #fee140 100%);
        }}
        .stat-card h3 {{
            margin: 0;
            font-size: 36px;
        }}
        .stat-card p {{
            margin: 10px 0 0 0;
            font-size: 14px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 30px;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 12px;
            text-align: left;
        }}
        th {{
            background-color: #4472C4;
            color: white;
            font-weight: bold;
        }}
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        tr:hover {{
            background-color: #f0f0f0;
        }}
        .badge {{
            display: inline-block;
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 12px;
            font-weight: bold;
        }}
        .badge.high {{
            background-color: #4CAF50;
            color: white;
        }}
        .badge.medium {{
            background-color: #FF9800;
            color: white;
        }}
        .badge.low {{
            background-color: #F44336;
            color: white;
        }}
        .badge.review {{
            background-color: #FF5722;
            color: white;
        }}
        .footer {{
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #ddd;
            text-align: center;
            color: #666;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📋 商标证书提取报告</h1>
        <p>生成时间: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}</p>

        <div class="summary">
            <div class="stat-card">
                <h3>{total}</h3>
                <p>总商标数</p>
            </div>
            <div class="stat-card green">
                <h3>{high_confidence}</h3>
                <p>高置信度</p>
            </div>
            <div class="stat-card orange">
                <h3>{needs_review}</h3>
                <p>需要复核</p>
            </div>
            <div class="stat-card red">
                <h3>{expiring_soon}</h3>
                <p>即将到期</p>
            </div>
        </div>

        <h2>详细信息</h2>
        <table>
            <thead>
                <tr>
                    <th>序号</th>
                    <th>注册人</th>
                    <th>注册号</th>
                    <th>国际分类</th>
                    <th>有效期限</th>
                    <th>置信度</th>
                    <th>到期状态</th>
                </tr>
            </thead>
            <tbody>
"""

        for i, item in enumerate(data, 1):
            confidence = item.get('overall_confidence', 0)
            confidence_level = item.get('confidence_level', 'unknown')

            if confidence_level == 'high':
                badge_class = 'high'
                badge_text = '高'
            elif confidence_level == 'medium':
                badge_class = 'medium'
                badge_text = '中'
            else:
                badge_class = 'low'
                badge_text = '低'

            html += f"""
                <tr>
                    <td>{i}</td>
                    <td>{item.get('注册人', '')}</td>
                    <td>{item.get('注册号', '')}</td>
                    <td>{item.get('国际分类', '')}</td>
                    <td>{item.get('有效期限', '')}</td>
                    <td><span class="badge {badge_class}">{badge_text} ({confidence:.0%})</span></td>
                    <td>{item.get('expiry_message', '')}</td>
                </tr>
"""

        html += """
            </tbody>
        </table>

        <div class="footer">
            <p>商标证书提取器 v2.0 - Enhanced with AI</p>
        </div>
    </div>
</body>
</html>
"""

        return html

    def export_to_pdf_report(self, data: List[Dict[str, Any]],
                            filename: str) -> Path:
        """Export to PDF report format"""
        # First generate HTML report
        html_path = self.export_to_html_report(data, filename + "_temp")

        # Convert HTML to PDF using weasyprint or pdfkit
        output_path = self.output_dir / f"{filename}.pdf"

        try:
            from weasyprint import HTML
            HTML(filename=str(html_path)).write_pdf(output_path)
            html_path.unlink()  # Delete temporary HTML file
            return output_path
        except ImportError:
            # weasyprint not available, try pdfkit
            try:
                import pdfkit
                pdfkit.from_file(str(html_path), str(output_path))
                html_path.unlink()
                return output_path
            except ImportError:
                # Both libraries unavailable, return HTML file
                print("Warning: weasyprint and pdfkit not available. Returning HTML report instead.")
                return html_path

    def export_all_formats(self, data: List[Dict[str, Any]],
                          base_filename: Optional[str] = None) -> Dict[ExportFormat, Path]:
        """
        Export to all supported formats

        Args:
            data: Trademark data
            base_filename: Base filename for all exports

        Returns:
            Dictionary mapping formats to output file paths
        """
        if base_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_filename = f"trademark_export_{timestamp}"

        results = {}

        for format in ExportFormat:
            try:
                output_path = self.export(data, format, base_filename)
                results[format] = output_path
                print(f"✓ Exported {format.value}: {output_path}")
            except Exception as e:
                print(f"✗ Failed to export {format.value}: {e}")

        return results


if __name__ == "__main__":
    # Test export manager
    test_data = [
        {
            '注册人': '阿里巴巴集团控股有限公司',
            '注册号': '12345678',
            '国际分类': '35',
            '有效期限': '2025-12-31',
            'overall_confidence': 0.95,
            'confidence_level': 'high',
            'requires_manual_review': False,
            'expiry_message': '有效（还有365天）',
            'requires_renewal': False,
        },
        {
            '注册人': '腾讯科技（深圳）有限公司',
            '注册号': '87654321',
            '国际分类': '9',
            '有效期限': '2024-06-30',
            'overall_confidence': 0.65,
            'confidence_level': 'medium',
            'requires_manual_review': True,
            'expiry_message': '⚠️ 警告：还有120天到期',
            'requires_renewal': True,
        },
    ]

    manager = ExportManager(output_dir=Path('/tmp/trademark_exports'))
    results = manager.export_all_formats(test_data)

    print("\nExported files:")
    for format, path in results.items():
        print(f"  {format.value}: {path}")
