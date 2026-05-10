#!/usr/bin/env python3
"""
Excel Processing Script
Processa planilhas Excel: extracao, validacao de formulas, criacao.
Implementa zero-error policy baseado no padrao Anthropic xlsx skill.

Usage:
    python process_xlsx.py input.xlsx --mode extract|validate|formulas|create
"""

import argparse
import json
import sys
import subprocess
import tempfile
from pathlib import Path
from typing import Optional


# Tipos de erro Excel
EXCEL_ERRORS = {
    "#VALUE!": "Tipo incompativel na formula",
    "#REF!": "Referencia invalida (celula deletada)",
    "#DIV/0!": "Divisao por zero",
    "#NAME?": "Nome de funcao ou range desconhecido",
    "#N/A": "Valor nao disponivel (lookup falhou)",
    "#NUM!": "Numero invalido",
    "#NULL!": "Intersecao de ranges nula",
    "#GETTING_DATA": "Dados externos sendo carregados",
    "#SPILL!": "Range de spill bloqueado",
    "#CALC!": "Erro de calculo (array)",
}


def check_dependencies() -> dict:
    """Verifica dependencias instaladas."""
    deps = {
        "openpyxl": False,
        "pandas": False,
        "libreoffice": False,
    }

    try:
        import openpyxl
        deps["openpyxl"] = True
    except ImportError:
        pass

    try:
        import pandas
        deps["pandas"] = True
    except ImportError:
        pass

    # Verificar LibreOffice
    try:
        result = subprocess.run(
            ["libreoffice", "--version"],
            capture_output=True,
            text=True,
            timeout=5,
        )
        if result.returncode == 0:
            deps["libreoffice"] = True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    return deps


def extract_data(xlsx_path: str) -> dict:
    """Extrai dados de todas as sheets."""
    import openpyxl

    result = {
        "source": xlsx_path,
        "sheets": [],
        "metadata": {},
    }

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    result["metadata"] = {
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            "name": sheet_name,
            "rows": ws.max_row,
            "columns": ws.max_column,
            "data": [],
        }

        # Extrair dados (primeiras 1000 linhas por seguranca)
        for row in ws.iter_rows(max_row=min(1000, ws.max_row)):
            row_data = []
            for cell in row:
                cell_value = {
                    "value": cell.value,
                    "coordinate": cell.coordinate,
                }
                # Se tem formula, incluir
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    cell_value["formula"] = cell.value
                row_data.append(cell_value)
            sheet_data["data"].append(row_data)

        result["sheets"].append(sheet_data)

    wb.close()
    return result


def list_formulas(xlsx_path: str) -> dict:
    """Lista todas as formulas na planilha."""
    import openpyxl

    result = {
        "source": xlsx_path,
        "formulas": [],
        "total_count": 0,
    }

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    result["formulas"].append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "formula": cell.value,
                    })

    result["total_count"] = len(result["formulas"])
    wb.close()
    return result


def validate_with_openpyxl(xlsx_path: str) -> dict:
    """Valida planilha usando openpyxl (checagem basica)."""
    import openpyxl

    result = {
        "source": xlsx_path,
        "status": "valid",
        "errors": [],
        "warnings": [],
    }

    # Carregar com data_only=True para ver valores calculados
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        result["status"] = "invalid"
        result["errors"].append({
            "type": "load_error",
            "message": str(e),
        })
        return result

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    # Verificar erros Excel
                    for error_code, description in EXCEL_ERRORS.items():
                        if value == error_code or value.startswith(error_code):
                            result["errors"].append({
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "type": error_code,
                                "description": description,
                            })

    if result["errors"]:
        result["status"] = "errors_found"
        result["total_errors"] = len(result["errors"])

    wb.close()
    return result


def validate_with_libreoffice(xlsx_path: str) -> dict:
    """
    Valida planilha usando LibreOffice headless (recalcula formulas).
    Baseado no padrao recalc.py do xlsx skill da Anthropic.
    """
    result = {
        "source": xlsx_path,
        "status": "valid",
        "validation_method": "libreoffice",
        "errors": [],
    }

    # Criar diretorio temporario
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        output_path = tmp_path / "recalculated.xlsx"

        # Macro para recalcular
        macro_code = '''
import uno
from com.sun.star.beans import PropertyValue

def recalculate():
    desktop = XSCRIPTCONTEXT.getDesktop()
    doc = desktop.getCurrentComponent()
    sheets = doc.getSheets()

    for i in range(sheets.getCount()):
        sheet = sheets.getByIndex(i)
        # Forcar recalculo
        sheet.calculateAll()

    return None
'''

        try:
            # Converter via LibreOffice (forca recalculo)
            cmd = [
                "libreoffice",
                "--headless",
                "--convert-to", "xlsx",
                "--outdir", str(tmp_path),
                xlsx_path,
            ]

            proc = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=60,
            )

            if proc.returncode != 0:
                result["status"] = "validation_failed"
                result["errors"].append({
                    "type": "libreoffice_error",
                    "message": proc.stderr,
                })
                return result

            # Verificar arquivo convertido
            converted_files = list(tmp_path.glob("*.xlsx"))
            if not converted_files:
                result["status"] = "validation_failed"
                result["errors"].append({
                    "type": "conversion_error",
                    "message": "LibreOffice nao gerou arquivo de saida",
                })
                return result

            # Validar arquivo recalculado
            recalc_result = validate_with_openpyxl(str(converted_files[0]))
            result["errors"] = recalc_result["errors"]
            result["status"] = recalc_result["status"]

        except subprocess.TimeoutExpired:
            result["status"] = "validation_failed"
            result["errors"].append({
                "type": "timeout",
                "message": "LibreOffice timeout (60s)",
            })
        except FileNotFoundError:
            result["status"] = "validation_failed"
            result["errors"].append({
                "type": "libreoffice_not_found",
                "message": "LibreOffice nao instalado",
                "install": "apt install libreoffice",
            })

    return result


def format_as_markdown(data: dict) -> str:
    """Formata dados extraidos como markdown."""
    lines = []
    lines.append(f"# Dados de: {data.get('source', 'unknown')}")
    lines.append("")

    for sheet in data.get("sheets", []):
        lines.append(f"## Sheet: {sheet.get('name', 'unnamed')}")
        lines.append(f"Linhas: {sheet.get('rows', 0)}, Colunas: {sheet.get('columns', 0)}")
        lines.append("")

        sheet_data = sheet.get("data", [])
        if sheet_data:
            # Header
            header = [cell.get("value", "") for cell in sheet_data[0]]
            lines.append("| " + " | ".join(str(h or "") for h in header) + " |")
            lines.append("| " + " | ".join("---" for _ in header) + " |")

            # Rows (primeiras 50)
            for row in sheet_data[1:51]:
                values = [cell.get("value", "") for cell in row]
                lines.append("| " + " | ".join(str(v or "") for v in values) + " |")

            if len(sheet_data) > 51:
                lines.append(f"_... e mais {len(sheet_data) - 51} linhas_")

        lines.append("")

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Processa planilhas Excel")
    parser.add_argument("input", nargs="?", help="Arquivo XLSX de entrada")
    parser.add_argument("--mode", choices=["extract", "validate", "formulas"],
                        default="extract", help="Modo de operacao")
    parser.add_argument("--output", choices=["json", "markdown"], default="json",
                        help="Formato de saida")
    parser.add_argument("--use-libreoffice", action="store_true",
                        help="Usar LibreOffice para validacao (mais preciso)")
    parser.add_argument("--check-deps", action="store_true",
                        help="Verificar dependencias")

    args = parser.parse_args()

    if args.check_deps:
        deps = check_dependencies()
        print(json.dumps(deps, indent=2))
        sys.exit(0 if deps["openpyxl"] else 1)

    if not args.input:
        parser.print_help()
        sys.exit(1)

    input_path = Path(args.input)
    if not input_path.exists():
        print(json.dumps({"error": f"Arquivo nao encontrado: {args.input}"}))
        sys.exit(1)

    deps = check_dependencies()
    if not deps["openpyxl"]:
        print(json.dumps({
            "error": "openpyxl nao instalado",
            "install": "pip install openpyxl",
        }))
        sys.exit(1)

    if args.mode == "extract":
        result = extract_data(str(input_path))
        if args.output == "markdown":
            print(format_as_markdown(result))
        else:
            print(json.dumps(result, indent=2, ensure_ascii=False, default=str))

    elif args.mode == "formulas":
        result = list_formulas(str(input_path))
        print(json.dumps(result, indent=2, ensure_ascii=False))

    elif args.mode == "validate":
        if args.use_libreoffice and deps["libreoffice"]:
            result = validate_with_libreoffice(str(input_path))
        else:
            result = validate_with_openpyxl(str(input_path))

        print(json.dumps(result, indent=2, ensure_ascii=False))

        # Exit code baseado no status
        if result["status"] == "errors_found":
            sys.exit(1)


if __name__ == "__main__":
    main()
