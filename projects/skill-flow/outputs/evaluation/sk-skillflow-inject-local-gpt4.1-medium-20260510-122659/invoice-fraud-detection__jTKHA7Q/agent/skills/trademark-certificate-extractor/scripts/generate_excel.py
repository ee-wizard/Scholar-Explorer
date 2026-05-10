#!/usr/bin/env python3
"""
Excel generation script for trademark certificate data

Creates Excel files from extracted trademark certificate information
with embedded trademark logo images.
"""

import sys
from pathlib import Path
from typing import List, Dict, Any, Optional
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from datetime import datetime
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_validity_period(date_str: str) -> datetime:
    """
    Parse validity period string to datetime object

    Supports multiple date formats:
    - YYYY-MM-DD
    - YYYY/MM/DD
    - DD/MM/YYYY
    - YYYY年MM月DD日
    - MM-DD-YYYY
    - YYYY.MM.DD

    Args:
        date_str: Date string to parse

    Returns:
        datetime object (defaults to epoch if parsing fails)
    """
    if not date_str:
        return datetime.min

    date_formats = [
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%d/%m/%Y',
        '%Y年%m月%d日',
        '%m-%d-%Y',
        '%d-%m-%Y',
        '%Y.%m.%d'
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except (ValueError, TypeError):
            continue

    # If all formats fail, return minimum datetime
    return datetime.min


def sort_trademark_data(data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Sort trademark data by:
    1. Registrant (alphabetical order)
    2. Validity period (ascending)

    Args:
        data: List of trademark dictionaries

    Returns:
        Sorted list of trademark dictionaries
    """
    def sort_key(entry):
        registrant = entry.get('注册人', '')
        validity_period = entry.get('有效期限', '')

        # Parse date for sorting
        date_obj = parse_validity_period(validity_period)

        # Sort key: (registrant, date_timestamp)
        return (
            registrant,  # Alphabetical order
            date_obj.timestamp() if date_obj != datetime.min else 0  # Ascending order
        )

    return sorted(data, key=sort_key)


def add_logo_to_cell(ws, row_num: int, col_num: int, logo_path: Path, max_height: float = 50.0, max_width: float = 150.0):
    """
    Add a logo image to an Excel cell

    Args:
        ws: Excel worksheet
        row_num: Row number (1-indexed)
        col_num: Column number (1-indexed)
        logo_path: Path to the logo image file
        max_height: Maximum height in points (default 50)
        max_width: Maximum width in points (default 150)
    """
    if not logo_path.exists():
        print(f"Warning: Logo file not found: {logo_path}", file=sys.stderr)
        return

    try:
        img = OpenpyxlImage(str(logo_path))

        # Scale image to fit within max dimensions while maintaining aspect ratio
        original_height = img.height
        original_width = img.width

        # Calculate scale ratios
        height_ratio = max_height / original_height
        width_ratio = max_width / original_width
        
        # Use the smaller ratio to ensure it fits both dimensions
        scale_ratio = min(height_ratio, width_ratio, 1.0) # Don't upscale
        
        img.height = original_height * scale_ratio
        img.width = original_width * scale_ratio

        # Position the image
        img.anchor = f"{get_column_letter(col_num)}{row_num}"

        # Add image to worksheet
        ws.add_image(img)

        # Adjust row height to accommodate the image (plus padding)
        current_height = ws.row_dimensions[row_num].height or 15
        ws.row_dimensions[row_num].height = max(max_height + 10, current_height)

    except Exception as e:
        print(f"Error adding logo to cell: {e}", file=sys.stderr)


def create_excel(data: List[Dict[str, Any]], output_path: Path) -> bool:
    """
    Create an Excel file with trademark certificate data and embedded logos

    Args:
        data: List of dictionaries containing trademark information
        output_path: Path for the output Excel file

    Returns:
        True if successful, False otherwise
    """
    try:
        # Sort data before creating Excel
        sorted_data = sort_trademark_data(data)

        wb = Workbook()
        ws = wb.active
        ws.title = "商标证书信息"

        # Define column headers
        headers = ['序号', '商标标识', '注册人', '注册号', '国际分类', '有效期限']

        # Write headers with styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Define border styles for grouping
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_top_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='medium'),
            bottom=Side(style='thin')
        )

        # Track previous registrant for grouping
        previous_registrant = None

        # Write data rows
        for row_num, entry in enumerate(sorted_data, 2):
            current_registrant = entry.get('注册人', '')

            # Use thick top border for new registrant groups
            border_style = thick_top_border if current_registrant != previous_registrant else thin_border

            # Write text data
            # Generate continuous serial number based on row index (row_num starts at 2)
            ws.cell(row=row_num, column=1, value=row_num - 1).border = border_style
            ws.cell(row=row_num, column=3, value=entry.get('注册人', '')).border = border_style
            ws.cell(row=row_num, column=4, value=entry.get('注册号', '')).border = border_style
            ws.cell(row=row_num, column=5, value=entry.get('国际分类', '')).border = border_style
            ws.cell(row=row_num, column=6, value=entry.get('有效期限', '')).border = border_style

            # Add logo image to column 2
            logo_path_str = entry.get('商标标识图片路径', '')
            if logo_path_str:
                logo_path = Path(logo_path_str)
                add_logo_to_cell(ws, row_num, 2, logo_path)
            else:
                ws.cell(row=row_num, column=2, value='').border = border_style

            previous_registrant = current_registrant

        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0

            # Check header length
            max_length = max(max_length, len(headers[col_num - 1]))

            # Check data length (skip logo column)
            if col_num != 2:  # Skip logo column
                for row_num in range(2, len(sorted_data) + 2):
                    cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
                    max_length = max(max_length, len(cell_value))

            # Set column width (with some padding)
            # Make logo column wider
            if col_num == 2:  # Logo column
                adjusted_width = 25
            else:
                adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Apply alignment to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # Save the workbook
        wb.save(output_path)
        return True

    except Exception as e:
        print(f"Error creating Excel file: {e}", file=sys.stderr)
        return False


def open_excel(file_path: Path):
    """
    Open the Excel file using the default system application

    Args:
        file_path: Path to the Excel file to open
    """
    import subprocess
    import platform

    try:
        if platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', str(file_path)], check=True)
        elif platform.system() == 'Windows':
            os.startfile(str(file_path))
        else:  # Linux
            subprocess.run(['xdg-open', str(file_path)], check=True)
    except Exception as e:
        print(f"Could not open Excel file automatically: {e}", file=sys.stderr)
        print(f"Please open the file manually: {file_path}")


def main():
    """Command-line interface for the Excel generation script"""
    if len(sys.argv) < 2:
        print("Usage: python generate_excel.py <output_path> [json_data]")
        print("Example: python generate_excel.py trademark_data.json")
        print("Example: python generate_excel.py trademark_data.json '[{\"序号\":\"1\",...}]'")
        sys.exit(1)

    # For testing with JSON input
    if len(sys.argv) == 3:
        import json
        try:
            data = json.loads(sys.argv[2])
        except json.JSONDecodeError:
            print("Invalid JSON data", file=sys.stderr)
            sys.exit(1)
    else:
        # Try to load JSON from file
        try:
            import json
            with open(sys.argv[1], 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Determine output filename based on Registrant
            registrant_name = "导出数据"
            if data and isinstance(data, list):
                for entry in data:
                    name = entry.get('注册人')
                    if name:
                        registrant_name = name
                        break
            
            # Sanitize filename
            invalid_chars = '<>:"/\\|?*'
            for char in invalid_chars:
                registrant_name = registrant_name.replace(char, '')
            registrant_name = registrant_name.strip()
            
            if not registrant_name:
                registrant_name = "导出数据"

            output_path = Path(sys.argv[1]).parent / f"{registrant_name}-商标注册信息.xlsx"
        except Exception as e:
            print(f"Error loading data: {e}", file=sys.stderr)
            # Example data for testing
            data = []
            output_path = Path(sys.argv[1])

    if create_excel(data, output_path):
        print(f"Excel file created successfully: {output_path}")
        
        # Auto-cleanup intermediate files
        try:
            # Add current script directory to path to import cleanup
            script_dir = Path(__file__).parent
            if str(script_dir) not in sys.path:
                sys.path.append(str(script_dir))
            
            from cleanup import cleanup
            
            working_dir = output_path.parent
            print(f"Automatically cleaning up intermediate files in {working_dir}...")
            cleanup(working_dir)
            
        except Exception as e:
            print(f"Warning: Auto-cleanup failed: {e}", file=sys.stderr)

        open_excel(output_path)
    else:
        print("Failed to create Excel file", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    import os
    main()
