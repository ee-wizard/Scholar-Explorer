#!/usr/bin/env python3
"""
Recalculate formulas and scan for errors in Excel files.

Usage:
    python recalc.py <file.xlsx>

Returns:
    JSON output with error counts and locations for each sheet
"""

import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


ERROR_TYPES = {
    '#DIV/0!': 'Division by zero',
    '#N/A': 'Value not available',
    '#NAME?': 'Unrecognized name in formula',
    '#NULL!': 'Null intersection',
    '#NUM!': 'Invalid numeric value',
    '#REF!': 'Invalid cell reference',
    '#VALUE!': 'Wrong value type',
}


def scan_sheet_for_errors(sheet):
    """Scan a worksheet for Excel error values."""
    errors = []
    error_count = 0

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in ERROR_TYPES:
                errors.append({
                    'cell': cell.coordinate,
                    'error': cell.value,
                    'description': ERROR_TYPES[cell.value]
                })
                error_count += 1

    return {
        'error_count': error_count,
        'errors': errors
    }


def recalc_and_scan(file_path):
    """Recalculate formulas and scan for errors in an Excel file."""
    try:
        # Load workbook with data_only=False to recalculate formulas
        wb = load_workbook(file_path, data_only=False)

        # Force recalculation by setting calc mode
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True

        results = {
            'file': str(file_path),
            'sheets': {},
            'total_errors': 0
        }

        # Scan each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_results = scan_sheet_for_errors(sheet)
            results['sheets'][sheet_name] = sheet_results
            results['total_errors'] += sheet_results['error_count']

        # Save to recalculate formulas
        wb.save(file_path)
        wb.close()

        return results

    except InvalidFileException:
        return {
            'error': f'Invalid Excel file: {file_path}',
            'file': str(file_path)
        }
    except FileNotFoundError:
        return {
            'error': f'File not found: {file_path}',
            'file': str(file_path)
        }
    except Exception as e:
        return {
            'error': f'Unexpected error: {str(e)}',
            'file': str(file_path)
        }


def main():
    """Main entry point for the script."""
    if len(sys.argv) != 2:
        print('Usage: python recalc.py <file.xlsx>', file=sys.stderr)
        sys.exit(1)

    file_path = Path(sys.argv[1])
    results = recalc_and_scan(file_path)

    # Print JSON results
    print(json.dumps(results, indent=2))

    # Exit with error code if errors found
    if 'error' in results:
        sys.exit(1)
    elif results.get('total_errors', 0) > 0:
        sys.exit(1)
    else:
        sys.exit(0)


if __name__ == '__main__':
    main()
